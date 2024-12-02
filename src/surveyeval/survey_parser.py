#  Copyright (c) 2023-24 Higher Bar AI, PBC
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

"""Utility functions for reading and parsing survey files."""

import os
import logging
import csv
import re
import importlib.resources as pkg_resources
from openpyxl import load_workbook
from ai_workflows.document_utilities import DocumentInterface
from ai_workflows.llm_utilities import LLMInterface

# initialize global resources
parser_logger = logging.getLogger(__name__)
empty_form_path = str(pkg_resources.files('surveyeval').joinpath('resources/EmptyForm.xlsx'))


class SurveyInterface:
    """Interface for interacting with surveys."""

    llm_interface: LLMInterface = None
    doc_interface: DocumentInterface = None

    PARSING_JOB: str = """Your job is to extract survey questions or form fields from the file's content, organized by module, and to return it all in a specific JSON format. More specifically:

* **Your job is to extract verbatim text:** In the JSON you return, only ever include text content, directly quoted without modification (without even changing capitalization), from the survey text you are supplied (i.e., never add or invent any text and never revise or rephrase any text).

* **Only respond with valid JSON that precisely follows the format specified below:** Your response should only include valid JSON and nothing else. If you cannot find any questions, simply return an empty modules list. If you cannot find any modules but you can find at least one question, include all questions in a single module.

* **Treat translations as separate questions:** If you see one or more translated versions of a question, include them as separate questions in the JSON you return."""
    PARSING_JSON_SPEC: str = """Return JSON text (and only JSON text) according to this precise format:

* `modules` (required, but can be an empty list): An array of module objects, one for each main section, category, or group within which a questionnaire or digital form's questions or form fields are located. Examples of modules include: "Demographics", "Health", "Household members", "Education", etc. All questions are located within modules, even if there is no name, title, or introductory text for the module.
  * Each module object contains:
    * `module_name` (optional): A string representing the short name or identifier of the survey module. Examples of module names: "demographics", "health", "hhmembers", "factors", etc.
    * `module_title` (optional): A string representing the longer title of the survey module. Examples of module titles: "Household demographics", "People living in the household", "Factors that influence take-up", etc.
    * `module_intro` (optional): A string representing the introductory text or instructions at the start of the module. For example: "Now we're going to ask some questions about the members of your household." 
    * `questions` (required, but can be an empty list): An array of question objects, one for each question or form field within the module. When a question is supplied in more than one language, each translation should be included as a separate question with the proper language name in the `language` field. Never translate between languages or otherwise alter the question text in any way.
      * Each question object contains:
        * `question_id` (optional): A string representing the numeric or alphanumeric identifier of the question (usually located just before or at the beginning of the question).
        * `question` (optional): A string representing the exact text of the question or form field, including any introductory text that provides context or explanation. Often follows a unique question ID of some sort, like "2.01." or "gender:". Should not include response options, which should be included in the `options` field, or extra enumerator or interviewer instructions (including interview probes), which should be included in the `instructions` field. Be careful: the same question might be asked in multiple languages, and each translation should be included as a separate question with the proper language name in the `language` field. Never translate between languages or otherwise alter the question text in any way.
        * `instructions` (optional): A string representing instructions or guidance on how to ask or answer the question, including enumerator or interviewer instructions. If the question includes a list of specific response options, do NOT include those in the instructions. However, if there is guidance as to how to fill out an open-ended numeric or text response, or guidance about how to choose among the options, include that guidance here.
        * `options` (optional): A string representing the list of specific response options for multiple-choice questions, including both the label and the internal value (if specified) for each option. For example, a 'Male' label might be coupled with an internal value of '1', 'M', or even 'male'. Separate response options with a space, three pipe symbols ('|||'), and another space, and, if there is an internal value, add a space, three # symbols ('###'), and the internal value at the end of the label. For example: 'Male ### 1 ||| Female ### 2' (codes included) or 'Male ||| Female' (no codes); 'Yes ### yes ||| No ### no', 'Yes ### 1 ||| No ### 0', 'Yes ### y ||| No ### n', or 'YES ||| NO'. Do NOT include fill-in-the-blank content here, only multiple-choice options. If the question is open-ended, leave this field blank.
        * `language` (optional): A string representing the primary language in which the question text is written.

Remember:
 
* Always list `options` with labels and values like 'label ### value ||| label ### value' and `options` without separate labels and values like 'Male ||| Female'.
* Never rephrase, change capitalization, or otherwise modify text directly quoted from the survey. Only re-organize the text into the JSON format specified above.
"""
    PARSING_JSON_SCHEMA: str = """{
  "$schema": "http://json-schema.org/draft-07/schema#",
  "definitions": {
    "Question": {
      "type": "object",
      "properties": {
        "question_id": {
          "type": ["string", "null"]
        },
        "question": {
          "type": ["string", "null"]
        },
        "instructions": {
          "type": ["string", "null"]
        },
        "options": {
          "type": ["string", "null"]
        },
        "language": {
          "type": ["string", "null"]
        }
      },
      "required": []
    },
    "Module": {
      "type": "object",
      "properties": {
        "module_name": {
          "type": ["string", "null"]
        },
        "module_title": {
          "type": ["string", "null"]
        },
        "module_intro": {
          "type": ["string", "null"]
        },
        "questions": {
          "type": ["array", "null"],
          "items": {
            "$ref": "#/definitions/Question"
          }
        }
      },
      "required": []
    }
  },
  "type": "object",
  "properties": {
    "modules": {
      "type": "array",
      "items": {
        "$ref": "#/definitions/Module"
      }
    }
  },
  "required": ["modules"]
}
"""

    def __init__(self, openai_api_key: str = None, openai_model: str = None,
                 temperature: float = 0.0, total_response_timeout_seconds: int = 600, number_of_retries: int = 2,
                 seconds_between_retries: int = 5, azure_api_key: str = None, azure_api_engine: str = None,
                 azure_api_base: str = None, azure_api_version: str = None, langsmith_api_key: str = None,
                 langsmith_project: str = 'surveyeval', langsmith_endpoint: str = 'https://api.smith.langchain.com',
                 json_retries: int = 2, anthropic_api_key: str = None, anthropic_model: str = None,
                 bedrock_model: str = None, bedrock_region: str = "us-east-1", bedrock_aws_profile: str = None,
                 max_tokens: int = 4096):
        """
        Initialize a new survey interface with an LLM to help parse survey contents. Must supply LLM parameters for
        OpenAI (direct or via Azure) or Anthropic (direct or via AWS Bedrock).

        :param openai_api_key: OpenAI API key for accessing the LLM. Default is None.
        :type openai_api_key: str
        :param openai_model: OpenAI model name. Default is None.
        :type openai_model: str
        :param temperature: Temperature setting for the LLM. Default is 0.0.
        :type temperature: float
        :param total_response_timeout_seconds: Timeout for LLM responses in seconds. Default is 600.
        :type total_response_timeout_seconds: int
        :param number_of_retries: Number of retries for LLM calls. Default is 2.
        :type number_of_retries: int
        :param seconds_between_retries: Seconds between retries for LLM calls. Default is 5.
        :type seconds_between_retries: int
        :param azure_api_key: API key for Azure LLM. Default is None.
        :type azure_api_key: str
        :param azure_api_engine: Azure API engine name (deployment name; assumed to be the same as the OpenAI model
          name). Default is None.
        :type azure_api_engine: str
        :param azure_api_base: Azure API base URL. Default is None.
        :type azure_api_base: str
        :param azure_api_version: Azure API version. Default is None.
        :type azure_api_version: str
        :param langsmith_api_key: API key for LangSmith. Default is None.
        :type langsmith_api_key: str
        :param langsmith_project: LangSmith project name. Default is 'ai_workflows'.
        :type langsmith_project: str
        :param langsmith_endpoint: LangSmith endpoint URL. Default is 'https://api.smith.langchain.com'.
        :type langsmith_endpoint: str
        :param json_retries: Number of automatic retries for invalid JSON responses. Default is 2.
        :type json_retries: int
        :param anthropic_api_key: API key for Anthropic. Default is None.
        :type anthropic_api_key: str
        :param anthropic_model: Anthropic model name. Default is None.
        :type anthropic_model: str
        :param bedrock_model: AWS Bedrock model name. Default is None.
        :type bedrock_model: str
        :param bedrock_region: AWS Bedrock region. Default is "us-east-1".
        :type bedrock_region: str
        :param bedrock_aws_profile: AWS profile for Bedrock access. Default is None.
        :type bedrock_aws_profile: str
        :param max_tokens: Maximum tokens for LLM responses. Default is 4096.
        :type max_tokens: int
        """

        # initialize LLM and document interfaces
        self.llm_interface = LLMInterface(openai_api_key=openai_api_key, openai_model=openai_model,
                                          temperature=temperature,
                                          total_response_timeout_seconds=total_response_timeout_seconds,
                                          number_of_retries=number_of_retries,
                                          seconds_between_retries=seconds_between_retries,
                                          azure_api_key=azure_api_key, azure_api_engine=azure_api_engine,
                                          azure_api_base=azure_api_base, azure_api_version=azure_api_version,
                                          langsmith_api_key=langsmith_api_key, langsmith_project=langsmith_project,
                                          langsmith_endpoint=langsmith_endpoint, json_retries=json_retries,
                                          anthropic_api_key=anthropic_api_key, anthropic_model=anthropic_model,
                                          bedrock_model=bedrock_model, bedrock_region=bedrock_region,
                                          bedrock_aws_profile=bedrock_aws_profile, max_tokens=max_tokens)
        self.doc_interface = DocumentInterface(llm_interface=self.llm_interface)

    def read_survey_contents(self, file_path: str, use_llm: bool = True) -> str | dict:
        """
        Read the raw contents of a survey file.

        :param file_path: Path to the survey file.
        :type file_path: str
        :param use_llm: Whether to use the LLM to read the survey contents. Default is True. (False will extract text
            from the file using local methods.)
        :type use_llm: bool
        :return: Raw contents of the survey file, either as a dict (XLSForm or REDCap) or as a Markdown string.
        :rtype: str | dict
        """

        # get file extension to start
        file_ext = os.path.splitext(file_path)[1].lower()

        # if it's a .xlsx file, it might be an XLSForm
        if file_ext == ".xlsx":
            # if we can parse as XLSForm, return the parsed survey
            parsed_survey = self._read_xlsform(file_path)
            if parsed_survey:
                return parsed_survey

        # if it's a .csv file, it might be a REDCap data dictionary
        if file_ext == ".csv":
            # if we can parse as a REDCap data dictionary, return the parsed survey
            parsed_survey = self._read_redcap(file_path)
            if parsed_survey:
                return parsed_survey

        # otherwise, convert the file to Markdown format and return that
        if use_llm:
            return self.doc_interface.convert_to_markdown(file_path)
        else:
            doc_interface_no_llm = DocumentInterface()
            return doc_interface_no_llm.convert_to_markdown(file_path)

    def parse_survey_contents(self, survey_contents: str | dict, survey_context: str = "", max_chunk_size: int = 3000,
                              min_chunk_size: int = 1000) -> dict:
        """
        Parse raw survey contents into structured data.

        :param survey_contents: Raw survey contents, typically as returned by read_survey_contents(). If a string, it
            should be in Markdown format. If a dict, it should already be in the parsed data structure.
        :type survey_contents: str | dict
        :param survey_context: Context for the survey contents, if any. Default is an empty string. Survey context can
            be used to provide additional information to the parser, such as the type of survey or the survey's purpose.
        :type survey_context: str
        :param max_chunk_size: Maximum chunk size for LLM processing. Default is 3000.
        :type max_chunk_size: int
        :param min_chunk_size: Minimum chunk size for LLM processing. Default is 1000.
        :type min_chunk_size: int
        :return: A dict with modules (a dict with questions organized by module).
        :rtype: dict
        """

        # if a dict is supplied, that means the data was read as structured, and we can return it straight away
        if isinstance(survey_contents, dict):
            return survey_contents

        # otherwise, convert the Markdown content to structured data
        full_context = "The file contains a survey instrument or digital form."
        if survey_context:
            full_context += f" Additional context: {survey_context}"
        parsed_data = self.doc_interface.markdown_to_json(markdown=survey_contents,
                                                          json_context=full_context,
                                                          json_job=self.PARSING_JOB,
                                                          json_output_spec=self.PARSING_JSON_SPEC,
                                                          json_output_schema=self.PARSING_JSON_SCHEMA,
                                                          max_chunk_size=max_chunk_size,
                                                          min_chunk_size=min_chunk_size)

        # organize questions by module and question ID
        modules_output = {}
        question_module = {}
        unknown_module_count = 0
        unknown_id_count = 0
        for result in parsed_data:
            for module in result["modules"]:
                # construct module name from whatever details we have, defaulting to auto-naming as necessary
                if module.get("module_name") and module.get("module_title"):
                    if module.get("module_name").lower() == module.get("module_title").lower():
                        module_key = module["module_name"]
                    else:
                        module_key = f"{module['module_name']} - {module['module_title']}"
                elif module.get("module_name"):
                    module_key = module["module_name"]
                elif module.get("module_title"):
                    module_key = module["module_title"]
                else:
                    unknown_module_count += 1
                    module_key = f"MODULE_{unknown_module_count}"

                # if the same module key already exists with different capitalization, just use the original
                lower_module_key = module_key.lower()
                matching_key = next((key for key in modules_output if key.lower() == lower_module_key), None)
                if matching_key:
                    module_key = matching_key

                # run through all questions in module
                for question in module.get("questions", []):
                    # process question, if any
                    if question.get("question") and question["question"].strip():
                        # get question ID, if available
                        question_id = question.get("question_id", "").strip()
                        if not question_id:
                            # if no question ID is provided, generate a unique ID
                            unknown_id_count += 1
                            question_id = f"question_{unknown_id_count}"

                        if question_id in question_module:
                            # if we've seen the question ID before, add this version to the same module as before
                            question_list = modules_output[question_module[question_id]].setdefault("questions", {})
                        else:
                            # otherwise, add it to the current module, adding the module to the output as needed
                            if module_key not in modules_output:
                                # add module to output, ignoring module title if same as the module name
                                modules_output[module_key] = {
                                    "module_name": module.get("module_name", ""),
                                    "module_title": module.get("module_title", "") if module.get(
                                        "module_title", "").lower() != module.get("module_name", "").lower() else "",
                                    "module_intro": module.get("module_intro", ""),
                                    "questions": {}
                                }
                            question_list = modules_output[module_key].setdefault("questions", {})
                            question_module[question_id] = module_key

                        # parse and organize options, if any
                        options = []
                        if question.get("options"):
                            option_strs = [option.strip() for option in question["options"].split("|||")]
                            for option_str in option_strs:
                                if "###" in option_str:
                                    option_parts = option_str.split("###")
                                    if len(option_parts) == 2:
                                        option_parts = [part.strip() for part in option_parts]
                                        options.append({"label": option_parts[0], "value": option_parts[1]})
                                    else:
                                        options.append({"label": option_str, "value": option_str})
                                else:
                                    options.append({"label": option_str, "value": option_str})

                        # add question, grouped by module and question ID
                        if question_id not in question_list:
                            question_list[question_id] = []
                        question_list[question_id].append({
                            "question": question["question"],
                            "language": question.get("language", ""),
                            "options": options,
                            "instructions": question.get("instructions", ""),
                        })

        # return cleaned-up version of the data
        return self._clean_data(modules_output)

    @staticmethod
    def _read_xlsform(file_path: str) -> dict:
        """
        Read an XLSForm file and return its contents as a dictionary.

        :param file_path: Path to the XLSForm file.
        :type file_path: str
        :return: Parsed survey contents (or empty dict if it wasn't an XLSForm).
        :rtype: dict
        """

        try:
            # load the workbook
            wb = load_workbook(file_path)

            # if there are survey, choices, and settings worksheets, assume it's an XLSForm
            if 'survey' in wb.sheetnames and 'choices' in wb.sheetnames and 'settings' in wb.sheetnames:
                # process the XLSForm
                survey_ws = wb['survey']
                survey_columns = SurveyInterface._get_columns_from_headers(survey_ws)
                survey_data = [row for row in survey_ws.values][1:]
                choices_ws = wb['choices']
                choices_columns = SurveyInterface._get_columns_from_headers(choices_ws)
                choices_data = [row for row in choices_ws.values][1:]
                settings_ws = wb['settings']
                settings_columns = SurveyInterface._get_columns_from_headers(settings_ws)

                # hack for naming flexibility: if there's "name" column but no "value" column, rename "name" to "value"
                if 'name' in choices_columns and 'value' not in choices_columns:
                    choices_columns['value'] = choices_columns['name']
                    del choices_columns['name']

                # read the default_language from row 2 of the settings worksheet
                default_language = settings_ws.cell(row=2, column=settings_columns['default_language'] + 1).value

                # zip through the survey sheet to create a list of other languages in the form
                other_languages = []
                for column in survey_columns:
                    match = re.match(r"^label:(.*)", column)
                    if match:
                        language = match.group(1).strip()
                        # only consider adding languages we haven't already added (no dups wanted)
                        if language not in other_languages:
                            # only add languages that also have translations in the choices sheet
                            if f"label:{language}" in choices_columns:
                                other_languages.append(language)

                # next, zip through the survey sheet to create a list of survey items
                output_data = {}
                group_stack = []
                groupless_count = 0
                last_group = ""
                for row in survey_data:
                    if row[survey_columns['type']] and row[survey_columns['name']]:
                        type_ = str(row[survey_columns['type']]).strip() \
                            if row[survey_columns['type']] is not None else ""
                        name = str(row[survey_columns['name']]).strip() \
                            if row[survey_columns['name']] is not None else ""
                        label = str(row[survey_columns['label']]).strip() \
                            if row[survey_columns['label']] is not None else ""
                        hint = str(row[survey_columns['hint']]).strip() \
                            if row[survey_columns['hint']] is not None else ""

                        if type_ == "begin group" or type_ == "begin repeat":
                            # new group, add to survey_data
                            output_data[name] = {
                                "module_name": name,
                                "module_title": label,
                                "module_intro": "",
                                "questions": {}
                            }
                            # and remember that we're in this group
                            group_stack.append(name)
                        elif type_ == "end group" or type_ == "end repeat":
                            # pop the group stack to exit current group
                            group_stack.pop()
                        elif label and type_ not in ["calculate", "calculate_here", "start", "end", "deviceid",
                                                     "simserial", "phonenumber", "subscriberid", "caseid",
                                                     "audio audit", "text audit", "speed violations count",
                                                     "speed violations list", "speed violations audit"]:
                            # if question has a label isn't an invisible type, figure out which module it belongs to
                            if group_stack:
                                # group is most recent item added to the stack
                                module = group_stack[-1]
                            elif last_group and last_group.startswith("nomodule_"):
                                # we have a groupless module open and going, so continue with that
                                module = last_group
                            else:
                                # start a new groupless module
                                groupless_count += 1
                                module = f"nomodule_{groupless_count}"
                                # add new module to survey_data
                                output_data[module] = {
                                    "module_name": module,
                                    "module_title": "",
                                    "module_intro": "",
                                    "questions": {}
                                }
                            # remember this as the last group we were in
                            last_group = module

                            # assemble dictionary of questions
                            questions = {name: []}

                            # add item for base language first
                            item = {
                                "question": label,
                                "language": default_language,
                                "options": [],
                                "instructions": hint
                            }
                            # if the type is "select_one" or "select_multiple", add the choices
                            match = re.match(r"^(select_one|select_multiple) (.+)$", type_)
                            if match:
                                list_name = match.group(2)
                                # loop through all rows on the choices sheet with a matching list_name
                                for choice in choices_data:
                                    if choice[choices_columns['list_name']] == list_name:
                                        # add options one by one
                                        item["options"].append({
                                            "label": str(choice[choices_columns['label']]).strip(),
                                            "value": str(choice[choices_columns['value']]).strip()
                                        })
                            questions[name].append(item)

                            # next add translations (if any)
                            if other_languages:
                                for language in other_languages:
                                    translated_label = str(row[survey_columns[f"label:{language}"]]).strip() \
                                        if (f"label:{language}" in survey_columns
                                            and row[survey_columns[f"label:{language}"]] is not None) else ""
                                    # only add translations with labels
                                    if translated_label:
                                        item = {
                                            "question": translated_label,
                                            "language": language,
                                            "options": [],
                                            "instructions": str(row[survey_columns[f"hint:{language}"]]).strip()
                                            if f"hint:{language}" in survey_columns
                                               and row[survey_columns[f"hint:{language}"]] is not None else ""
                                        }
                                        # if the type is "select_one" or "select_multiple", add the choices
                                        match = re.match(r"^(select_one|select_multiple) (.+)$", type_)
                                        if match:
                                            list_name = match.group(2)
                                            # loop through all rows on the choices sheet with a matching list_name
                                            for choice in choices_data:
                                                if choice[choices_columns['list_name']] == list_name:
                                                    # add options one by one
                                                    choice_label = str(
                                                        choice[choices_columns[f"label:{language}"]]).strip() \
                                                        if (f"label:{language}" in choices_columns and
                                                            choice[choices_columns[f"label:{language}"]] is not None) \
                                                        else ""
                                                    if not choice_label:
                                                        # use default language label if translated label missing
                                                        choice_label = str(choice[choices_columns['label']]).strip()
                                                    item["options"].append({
                                                        "label": choice_label,
                                                        "value": str(choice[choices_columns['value']]).strip()
                                                        if f"hint:{language}" in choices_columns
                                                           and choice[choices_columns['value']] is not None else ""
                                                    })
                                        questions[name].append(item)

                            # update our output data with the new questions (translations for the current field)
                            output_data[module]["questions"].update(questions)

                # return structured data
                return SurveyInterface._clean_data(output_data)
        except Exception as e:
            parser_logger.log(logging.WARNING, f"Error trying to read {file_path} as XLSForm: {e}")

        # if we didn't successfully parse as XLSForm, return empty dict
        return {}

    @staticmethod
    def _read_redcap(file_path: str) -> dict:
        """
        Read a REDCap data dictionary and return its contents as a dictionary.

        :param file_path: Path to the REDCap data dictionary.
        :type file_path: str
        :return: Parsed survey contents (or empty dict if it wasn't a REDCap data dictionary).
        :rtype: dict
        """

        try:
            with open(file_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)

                # assume REDCap data dictionary format if "Field Type" column is present
                if 'Field Type' in reader.fieldnames:
                    # process REDCap data dictionary
                    all_questions = {}
                    for row in reader:
                        assert isinstance(row, dict)  # (type hint for IDE)
                        field_type = row['Field Type']
                        # only read supported field types
                        if field_type in ['descriptive', 'text', 'number', 'radio', 'checkboxes', 'dropdown', 'slider',
                                          'yesno', 'truefalse']:
                            question_id = row['Variable / Field Name']
                            options = []
                            if field_type in ['radio', 'checkboxes', 'dropdown']:
                                # parse choices from "value, label | value, label | ..." pairs
                                choice_strs = row.get('Choices, Calculations, OR Slider Labels', '').split('|')
                                for choice_str in choice_strs:
                                    value, label = choice_str.split(',', 1)
                                    value = value.strip()
                                    label = label.strip()
                                    options.append({'label': label, 'value': value})
                            elif field_type == 'slider':
                                # parse choices from "label | label | ..." format
                                slider_strs = row.get('Choices, Calculations, OR Slider Labels', '').split('|')
                                for slider_str in slider_strs:
                                    options.append({'label': slider_str, 'value': slider_str})
                            elif field_type == 'yesno':
                                options.append({'label': "yes", 'value': "1"})
                                options.append({'label': "no", 'value': "0"})
                            elif field_type == 'truefalse':
                                options.append({'label': "true", 'value': "1"})
                                options.append({'label': "false", 'value': "0"})

                            all_questions[question_id] = [
                                {
                                    'question': row['Field Label'],
                                    'instructions': row.get('Field Note', ''),
                                    'options': options,
                                    'language': "Unknown"  # (language not specified in REDCap data dictionary)
                                }
                            ]

                        # return all questions in a single module
                        form_data = {
                            "REDCap_module": {
                                "module_name": "REDCap_module",
                                "module_title": "",
                                "module_intro": "",
                                "questions": all_questions
                            }
                        }
                        return form_data
        except Exception as e:
            parser_logger.log(logging.WARNING, f"Error trying to read {file_path} as REDCap data dictionary: {e}")

        # if we didn't successfully parse as a REDCap data dictionary, return empty dict
        return {}

    @staticmethod
    def _clean_data(data: dict) -> dict:
        """
        Clean the provided data by removing empty modules and questions — and by removing duplicate questions,
        keeping the ones with more content.

        :param data: The data to be cleaned.
        :type data: dict
        :return: The cleaned data.
        :rtype: dict
        """

        cleaned_data = {}
        # run through all modules
        for key, dt in data.items():
            # assemble module questions, dropping empty questions
            questions = {question: question_data for question, question_data in dt['questions'].items()
                         if question and question_data}

            # skip empty modules
            if not questions:
                parser_logger.log(logging.INFO, f"Skipping empty module: {key}")
            else:
                parser_logger.log(logging.INFO, f"* Module: {key}")
                # for each question, look for questions with the same language
                for question, question_data in questions.items():
                    if question_data:
                        to_remove = set()
                        for i in range(len(question_data)):
                            for j in range(i + 1, len(question_data)):
                                if question_data[i]['language'] == question_data[j]['language']:
                                    # compare total string lengths and mark the shorter one for removal
                                    length_i = SurveyInterface._total_string_length(question_data[i])
                                    length_j = SurveyInterface._total_string_length(question_data[j])
                                    if length_i > length_j:
                                        to_remove.add(j)
                                        parser_logger.log(logging.INFO, f"  * Dropping duplicate question with "
                                                          f"shorter content: "
                                                          f"{question} - {question_data[j]['language']} - "
                                                          f"{question_data[j]['question']} "
                                                          f"({length_j} <= {length_i})")
                                    else:
                                        to_remove.add(i)
                                        parser_logger.log(logging.INFO, f"  * Dropping duplicate question with "
                                                          f"shorter content: "
                                                          f"{question} - {question_data[i]['language']} - "
                                                          f"{question_data[i]['question']} "
                                                          f"({length_i} <= {length_j})")

                        # remove duplicates after identifying them
                        for index in sorted(to_remove, reverse=True):
                            question_data.pop(index)

                # add module to cleaned data
                cleaned_data[key] = {
                    "module_name": dt['module_name'],
                    "module_title": dt['module_title'],
                    "module_intro": dt['module_intro'],
                    "questions": questions
                }

        return cleaned_data

    @staticmethod
    def output_parsed_data_to_xlsform(data: dict, form_id: str, form_title: str, output_file: str):
        """
        Output parsed data to an XLSForm file.

        :param data: Parsed data to output.
        :type data: dict
        :param form_id: Form ID to set in the XLSForm file.
        :type form_id: str
        :param form_title: Form title to set in the XLSForm file.
        :type form_title: str
        :param output_file: Path to the output XLSForm file.
        :type output_file: str
        """

        # load the workbook
        wb = load_workbook(empty_form_path)
        survey_ws = wb['survey']
        survey_columns = SurveyInterface._get_columns_from_headers(survey_ws, 1)
        choices_ws = wb['choices']
        choices_columns = SurveyInterface._get_columns_from_headers(choices_ws, 1)
        settings_ws = wb['settings']
        settings_columns = SurveyInterface._get_columns_from_headers(settings_ws, 1)

        # set the 'form_title' and 'form_id' values in row 2
        settings_ws.cell(row=2, column=settings_columns['form_id'], value=form_id)
        settings_ws.cell(row=2, column=settings_columns['form_title'], value=form_title)

        # initialize our row counters to start adding at the second row
        survey_row_counter = 2
        choices_row_counter = 2

        # iterate over each module in the data
        primary_language = ""
        for module in data.values():
            # create a safe version of the module name by replacing anything not a digit or a letter with an _
            safe_module_name = re.sub(r'\W+', '_', module['module_name']).lower()

            # open a group for the module
            survey_ws.cell(row=survey_row_counter, column=survey_columns['type'], value='begin group')
            survey_ws.cell(row=survey_row_counter, column=survey_columns['name'], value=safe_module_name)
            if module['module_title']:
                survey_ws.cell(row=survey_row_counter, column=survey_columns['label'], value=module['module_title'])
            survey_row_counter += 1

            # if there is a 'module_intro' value in the module, add it as a note field
            if module['module_intro']:
                survey_ws.cell(row=survey_row_counter, column=survey_columns['type'], value='note')
                survey_ws.cell(row=survey_row_counter, column=survey_columns['label'], value=module['module_intro'])
                survey_row_counter += 1

            # iterate over each question in the module
            for question_id, question_data in module['questions'].items():
                # create a safe version of the question ID by replacing anything not a digit or a letter with an _
                safe_question_id = re.sub(r'\W+', '_', question_id).lower()
                # if safe_question_id doesn't begin with a letter, put a 'q' on the front
                if not safe_question_id[0].isalpha():
                    safe_question_id = 'q' + safe_question_id

                # iterate over each translation for the question
                has_choices = False
                for translation in question_data:
                    # if we don't have our primary language yet, use the first language we find
                    if not primary_language:
                        primary_language = translation['language']
                        settings_ws.cell(row=2, column=settings_columns['default_language'], value=primary_language)

                    # set language suffix
                    if translation['language'] and translation['language'].lower() != primary_language.lower():
                        language_suffix = f":{translation['language']}"
                        label_column = 'label' + language_suffix
                        hint_column = 'hint' + language_suffix
                        # also add translation columns as necessary
                        if label_column not in survey_columns:
                            SurveyInterface._add_header_to_first_empty_cell(survey_ws, label_column)
                            survey_columns = SurveyInterface._get_columns_from_headers(survey_ws, 1)
                        if translation['options'] and label_column not in choices_columns:
                            SurveyInterface._add_header_to_first_empty_cell(choices_ws, label_column)
                            choices_columns = SurveyInterface._get_columns_from_headers(choices_ws, 1)
                        if hint_column not in survey_columns:
                            SurveyInterface._add_header_to_first_empty_cell(survey_ws, hint_column)
                            survey_columns = SurveyInterface._get_columns_from_headers(survey_ws, 1)
                    else:
                        label_column = 'label'
                        hint_column = 'hint'
                        # if there are options, we'll need to add them to the choices sheet (all at the end)
                        #   (we let the primary language govern whether there are choices)
                        if translation['options']:
                            has_choices = True

                    # add the question to the survey sheet
                    survey_ws.cell(row=survey_row_counter, column=survey_columns['type'],
                                   value='text' if not has_choices else 'select_one ' + safe_question_id)
                    survey_ws.cell(row=survey_row_counter, column=survey_columns['name'], value=safe_question_id)
                    survey_ws.cell(row=survey_row_counter, column=survey_columns[label_column],
                                   value=translation['question'])
                    survey_ws.cell(row=survey_row_counter, column=survey_columns[hint_column],
                                   value=translation['instructions'])

                # increment survey row when we're done adding all translations
                survey_row_counter += 1

                # then, finally, add all choice options (with all translations)
                if has_choices:
                    values_added = []
                    for translation in question_data:
                        for option in translation['options']:
                            if option['value'] not in values_added:
                                # if we haven't added this option value yet, add it for all translations
                                choices_ws.cell(row=choices_row_counter, column=choices_columns['list_name'],
                                                value=safe_question_id)
                                choices_ws.cell(row=choices_row_counter, column=choices_columns['value'],
                                                value=option['value'])
                                for inner_translation in question_data:
                                    for inner_option in inner_translation['options']:
                                        if inner_option['value'] == option['value']:
                                            if (inner_translation['language'] and
                                                    inner_translation['language'].lower() != primary_language.lower()):
                                                label_column = 'label' + f":{inner_translation['language']}"
                                            else:
                                                label_column = 'label'
                                            choices_ws.cell(row=choices_row_counter,
                                                            column=choices_columns[label_column],
                                                            value=inner_option['label'])
                                choices_row_counter += 1
                                values_added.append(option['value'])

            # close the module's group
            survey_ws.cell(row=survey_row_counter, column=survey_columns['type'], value='end group')
            survey_ws.cell(row=survey_row_counter, column=survey_columns['name'], value=safe_module_name)
            survey_row_counter += 1

        # Save the workbook to the specified output file
        wb.save(output_file)

    @staticmethod
    def _add_header_to_first_empty_cell(worksheet, header):
        for cell in worksheet[1]:
            if cell.value is None or not cell.value:
                cell.value = header
                break
        else:  # no break, meaning no empty cell was found
            max_column = worksheet.max_column
            worksheet.cell(row=1, column=max_column + 1, value=header)

    @staticmethod
    def _get_columns_from_headers(worksheet, index_boost: int = 0) -> dict:
        return {cell.value: i + index_boost for i, cell in enumerate(worksheet[1])
                if cell.value is not None and cell.value.strip() != ''}

    @staticmethod
    def _total_string_length(d: dict) -> int:
        # sum length of all strings, including those in nested lists of dicts
        total_length = 0
        for value in d.values():
            if isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        total_length += SurveyInterface._total_string_length(item)
            else:
                total_length += len(str(value))

        return total_length
